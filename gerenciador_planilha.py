import os
import string
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
from openpyxl.utils import get_column_letter

# Tentativa de importar psutil (opcional, para verificar planilha aberta)
try:
    import psutil
except ImportError:
    psutil = None
    print("⚠️ psutil não instalado. Verificação de planilha aberta limitada.")


class GerenciadorPlanilha:
    
    # ==================== MAPEAMENTO DE COLUNAS ====================
    # Datas por trimestre (para frequência e planejamento)
    DATAS_POR_TRIMESTRE = {
        1: {'inicio': 3, 'fim': 73},   # 1º TRI: colunas C a BS
        2: {'inicio': 5, 'fim': 75},   # 2º TRI: colunas E a BW (PULA C e D)
        3: {'inicio': 7, 'fim': 77}    # 3º TRI: colunas G a BY (PULA C,D,E,F)
    }
    
    # Colunas de PLANEJAMENTO (data e tema) por trimestre
    PLANEJAMENTO_COLUNAS = {
        1: {'data': 95, 'tema': 96},   # 1º Trimestre: CQ(95)=data, CR(96)=tema
        2: {'data': 100, 'tema': 101}, # 2º Trimestre
        3: {'data': 103, 'tema': 104}  # 3º Trimestre
    }
    
    # Colunas travadas (com fórmulas) - NÃO ESCREVER!
    COLUNAS_TRAVADAS = {
        1: [80],      # 1º Trimestre
        2: [85],      # 2º Trimestre
        3: [87]       # 3º Trimestre
    }
    
    def __init__(self, caminho):
        """
        Inicializa o gerenciador de planilha
        
        Args:
            caminho (str): Caminho completo do arquivo Excel
        """
        self.caminho = caminho
        
        # Validação prévia
        if not os.path.exists(caminho):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
        
        try:
            # data_only=False para preservar fórmulas
            self.wb = openpyxl.load_workbook(self.caminho, data_only=False)
        except Exception as e:
            if "Content_Types" in str(e):
                raise Exception(f"Arquivo Excel corrompido ou inválido. Abra e salve novamente no Excel.\nErro: {e}")
            else:
                raise
        
        self.linha_nomes_inicio = 9
        self.linha_nomes_fim = 45
        
        # Cache de alunos: {nome_upper: linha}
        self.cache_alunos = {}
        
        # Carrega os alunos do 1º trimestre (base principal)
        self._carregar_cache_alunos()
        
        # Verifica consistência com outros trimestres
        self._validar_cache_alunos()
    
    def _carregar_cache_alunos(self):
        """
        Carrega os nomes dos alunos do 1º trimestre para o cache.
        Usa um workbook temporário com data_only=True para avaliar as fórmulas.
        """
        try:
            # Abre temporariamente com data_only=True para ler os valores das fórmulas
            wb_temp = openpyxl.load_workbook(self.caminho, data_only=True)
            aba_base = "1º TRIMESTRE"
            
            if aba_base not in wb_temp.sheetnames:
                print(f"⚠️ Aviso: Aba '{aba_base}' não encontrada. Cache de alunos vazio.")
                wb_temp.close()
                return
            
            ws_temp = wb_temp[aba_base]
            
            # Escaneia a coluna B (nomes dos alunos)
            for linha in range(self.linha_nomes_inicio, self.linha_nomes_fim + 1):
                celula_nome = ws_temp.cell(row=linha, column=2)
                nome_aluno = celula_nome.value
                
                if nome_aluno and str(nome_aluno).strip():
                    nome_upper = str(nome_aluno).strip().upper()
                    self.cache_alunos[nome_upper] = linha
                    print(f"  📌 Aluno cacheado: {nome_aluno} (linha {linha})")
            
            wb_temp.close()
            print(f"✅ Cache carregado com {len(self.cache_alunos)} alunos do 1º trimestre")
            
        except Exception as e:
            print(f"⚠️ Erro ao carregar cache de alunos: {e}")
            self.cache_alunos = {}
    
    def _validar_cache_alunos(self):
        """
        Valida se os outros trimestres têm a mesma lista de alunos.
        Apenas um warning, não impede o funcionamento.
        """
        try:
            wb_temp = openpyxl.load_workbook(self.caminho, data_only=True)
            
            for trimestre in [2, 3]:
                nome_aba = f"{trimestre}º TRIMESTRE"
                if nome_aba not in wb_temp.sheetnames:
                    continue
                
                ws_temp = wb_temp[nome_aba]
                alunos_aba = set()
                
                for linha in range(self.linha_nomes_inicio, self.linha_nomes_fim + 1):
                    nome = ws_temp.cell(row=linha, column=2).value
                    if nome and str(nome).strip():
                        alunos_aba.add(str(nome).strip().upper())
                
                if len(alunos_aba) != len(self.cache_alunos):
                    print(f"⚠️ Atenção: Aba '{nome_aba}' tem {len(alunos_aba)} alunos, mas o 1º trimestre tem {len(self.cache_alunos)}")
            
            wb_temp.close()
            
        except Exception as e:
            print(f"⚠️ Não foi possível validar cache: {e}")
    
    def recarregar_cache_alunos(self):
        """
        Recarrega o cache de alunos (útil se a lista de alunos foi alterada)
        """
        print("🔄 Recarregando cache de alunos...")
        self.cache_alunos = {}
        self._carregar_cache_alunos()
        self._validar_cache_alunos()
    
    def _localizar_linha_aluno(self, ws, nome_aluno):
        """
        Localiza a linha do aluno usando o cache em memória.
        
        Args:
            ws: Worksheet do Excel (não usado, mantido para compatibilidade)
            nome_aluno (str): Nome do aluno
            
        Returns:
            int or None: Número da linha ou None se não encontrado
        """
        nome_busca = str(nome_aluno).strip().upper()
        
        if nome_busca in self.cache_alunos:
            return self.cache_alunos[nome_busca]
        
        # Busca parcial (útil para nomes com acentos ou diferenças)
        for nome_cacheado, linha in self.cache_alunos.items():
            if nome_busca in nome_cacheado or nome_cacheado in nome_busca:
                print(f"  🔍 Aluno encontrado por similaridade: '{nome_aluno}' -> '{nome_cacheado}'")
                return linha
        
        print(f"  ❌ Aluno '{nome_aluno}' não encontrado no cache")
        return None
    
    def obter_lista_alunos(self):
        """Retorna a lista de todos os alunos do cache"""
        return list(self.cache_alunos.keys())
    
    def obter_quantidade_alunos(self):
        """Retorna a quantidade de alunos no cache"""
        return len(self.cache_alunos)

    def _get_column_letter(self, col_num):
        """Converte número da coluna para letra (ex: 1->A, 27->AA)"""
        return get_column_letter(col_num)

    def _get_mes_abreviado(self, num_mes):
        """Retorna abreviatura do mês em português (3 letras)"""
        meses = {
            1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
            5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
            9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
        }
        return meses.get(num_mes, "")

    def _formatar_data_linha7(self, data_obj):
        """
        Formata a data para ser exibida na LINHA 7 da planilha (formato VERTICAL)
        Exemplo: 09/02/2026 -> "F\nE\nV\n\n0\n9"
        
        Args:
            data_obj (datetime): Objeto de data
            
        Returns:
            str: Data formatada com quebras de linha
        """
        dia = data_obj.day
        mes_num = data_obj.month
        nome_mes = self._get_mes_abreviado(mes_num)
        
        dezena = dia // 10
        unidade = dia % 10
        
        # Empilha as letras do mês de trás para frente
        linhas = []
        for letra in reversed(nome_mes):
            linhas.append(letra)
        linhas.append("")  # linha em branco
        linhas.append(str(unidade))
        linhas.append(str(dezena))
        
        return '\n'.join(linhas)
    
    def _formatar_data_horizontal(self, data_obj, incluir_ano=True):
        """
        Formata a data para escrita HORIZONTAL (colunas de planejamento)
        
        Args:
            data_obj (datetime): Objeto de data
            incluir_ano (bool): Se True, formato dd/mm/aaaa; se False, dd/mm
        
        Returns:
            str: Data formatada horizontalmente
        """
        if incluir_ano:
            return data_obj.strftime("%d/%m/%Y")
        else:
            return data_obj.strftime("%d/%m")
    
    def _formatar_data_planejamento(self, data_obj):
        """
        Formata a data especificamente para as colunas de planejamento
        Formato: dd/mm (horizontal)
        
        Args:
            data_obj (datetime): Objeto de data
            
        Returns:
            str: Data formatada como dd/mm
        """
        return data_obj.strftime("%d/%m")

    def _celula_tem_formula(self, celula):
        """Verifica se uma célula tem fórmula"""
        if celula.data_type == 'f':
            return True
        if celula.value and isinstance(celula.value, str) and celula.value.startswith('='):
            return True
        return False

    def _pode_escrever_na_celula(self, celula, coluna, linha):
        """
        Verifica se a célula pode receber escrita
        
        Critérios:
        - Não pode ser célula mesclada
        - Não pode ter fórmula
        """
        if isinstance(celula, MergedCell):
            return False
        
        if self._celula_tem_formula(celula):
            return False
        
        return True

    def _obter_aba(self, num_trimestre):
        """Obtém a aba do trimestre especificado"""
        nome_aba = f"{num_trimestre}º TRIMESTRE"
        if nome_aba in self.wb.sheetnames:
            return self.wb[nome_aba]
        return None

    def limpar_frequencia_trimestre(self, num_trimestre):
        """
        Limpa todas as frequências do trimestre especificado (apenas células editáveis)
        """
        try:
            ws = self._obter_aba(num_trimestre)
            if not ws: 
                return False
            
            dados = self.DATAS_POR_TRIMESTRE.get(num_trimestre, {'inicio': 3, 'fim': 74})
            inicio, fim = dados['inicio'], dados['fim']
            
            limpos = 0
            for col in range(inicio, fim + 1):
                # Limpa data da linha 7
                celula_data = ws.cell(row=7, column=col)
                if self._pode_escrever_na_celula(celula_data, col, 7):
                    celula_data.value = None
                    limpos += 1
                    
                # Limpa faltas dos alunos
                for linha in range(self.linha_nomes_inicio, self.linha_nomes_fim + 1):
                    celula_falta = ws.cell(row=linha, column=col)
                    if self._pode_escrever_na_celula(celula_falta, col, linha):
                        celula_falta.value = None
                        limpos += 1
            
            print(f"✅ Limpeza concluída: {limpos} células editáveis limpas")
            return True
        except Exception as e:
            raise Exception(f"Erro ao limpar: {str(e)}")

    def lancar_nota(self, num_trimestre, nome_aluno, valor, tipo, limpar_antes=False):
        """
        Lança nota do aluno na planilha
        """
        try:
            tri_str = str(num_trimestre)
            ws = self._obter_aba(tri_str)
            if not ws: 
                print(f"❌ Aba do {num_trimestre}º Trimestre não encontrada!")
                return False
                
            linha = self._localizar_linha_aluno(ws, nome_aluno)
            if not linha:
                print(f"❌ Aluno '{nome_aluno}' não encontrado!")
                return False
            
            mapa_trimestres = {
                "1": {
                    "normal": {"inicio": 74, "fim": 78},
                    "recup": {"inicio": 81, "fim": 83},
                    "bloqueadas": self.COLUNAS_TRAVADAS.get(1, [])
                },
                "2": {
                    "normal": {"inicio": 79, "fim": 84},
                    "recup": {"inicio": 86, "fim": 88},
                    "bloqueadas": self.COLUNAS_TRAVADAS.get(2, [])
                },
                "3": {
                    "normal": {"inicio": 81, "fim": 86},
                    "recup": {"inicio": 88, "fim": 90},
                    "bloqueadas": self.COLUNAS_TRAVADAS.get(3, [])
                }
            }

            config = mapa_trimestres.get(tri_str)
            if not config:
                print(f"❌ Trimestre {num_trimestre} inválido!")
                return False
            
            if tipo == "normal":
                col_inicio = config["normal"]["inicio"]
                col_fim = config["normal"]["fim"]
            else:
                col_inicio = config["recup"]["inicio"]
                col_fim = config["recup"]["fim"]
            
            bloqueadas = config["bloqueadas"]

            if limpar_antes:
                for c in range(col_inicio, col_fim + 1):
                    if c not in bloqueadas:
                        celula = ws.cell(row=linha, column=c)
                        if self._pode_escrever_na_celula(celula, c, linha):
                            celula.value = None

            for c in range(col_inicio, col_fim + 1):
                if c in bloqueadas:
                    continue
                    
                celula = ws.cell(row=linha, column=c)
                if not self._pode_escrever_na_celula(celula, c, linha):
                    print(f"⚠️ Célula com fórmula na coluna {c}, ignorada!")
                    continue
                    
                if celula.value in [None, ""]:
                    celula.value = float(valor)
                    celula.alignment = Alignment(horizontal='center')
                    celula.font = Font(bold=True)
                    print(f"✅ Nota {valor} lançada na coluna {self._get_column_letter(c)}")
                    return True
                        
            print(f"⚠️ Sem espaço para nota do {tipo} no {num_trimestre}º trimestre")
            return False
            
        except Exception as e:
            print(f"❌ Erro na nota: {str(e)}")
            return False

    def lancar_frequencia_completa(self, nome_aba, coluna_alvo, status_alunos, limpar_antes=True):
        """
        Lança APENAS faltas (F). Presenças mantêm o padrão da planilha (vazio).
        Usa o cache de alunos para localizar as linhas rapidamente.
        
        Args:
            nome_aba (str): Nome da aba (ex: "1º TRIMESTRE")
            coluna_alvo (int): Número da coluna onde escrever as faltas
            status_alunos (dict): Dicionário {nome_aluno: "F" ou ""}
            limpar_antes (bool): Se deve limpar faltas existentes antes
        """
        try:
            print(f"\n🔍 lancar_frequencia_completa - nome_aba: {nome_aba}, coluna: {coluna_alvo}")
            
            if nome_aba not in self.wb.sheetnames:
                print(f"❌ Aba '{nome_aba}' não encontrada!")
                return False
                
            ws = self.wb[nome_aba]
            
            # Remove "F" antigos de células editáveis
            if limpar_antes:
                print(f"🗑️ Removendo 'F' antigos da coluna {coluna_alvo}...")
                for linha in set(self.cache_alunos.values()):
                    celula = ws.cell(row=linha, column=coluna_alvo)
                    if self._pode_escrever_na_celula(celula, coluna_alvo, linha):
                        if celula.value == "F":
                            celula.value = None
            
            # Escreve APENAS as faltas usando o cache
            lancamentos = 0
            for nome_aluno, simbolo in status_alunos.items():
                if simbolo == "F":
                    nome_upper = nome_aluno.strip().upper()
                    if nome_upper in self.cache_alunos:
                        linha = self.cache_alunos[nome_upper]
                        celula_final = ws.cell(row=linha, column=coluna_alvo)
                        if self._pode_escrever_na_celula(celula_final, coluna_alvo, linha):
                            celula_final.number_format = '@'
                            celula_final.value = "F"
                            celula_final.alignment = Alignment(horizontal='center')
                            lancamentos += 1
                            print(f"  ✅ Falta marcada: {nome_aluno}")
                    else:
                        print(f"  ⚠️ Aluno não encontrado no cache: {nome_aluno}")
            
            if lancamentos == 0:
                print("ℹ️ Nenhuma falta registrada (todos presentes)")
            else:
                print(f"✅ {lancamentos} falta(s) registradas")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao lançar frequência: {e}")
            import traceback
            traceback.print_exc()
            return False

    def buscar_coluna_por_data(self, nome_aba, data_alvo):
        """
        Busca a coluna que corresponde à data especificada (baseado na linha 7)
        
        Args:
            nome_aba (str): Nome da aba
            data_alvo (str ou datetime): Data a ser buscada
            
        Returns:
            int or None: Número da coluna ou None se não encontrada
        """
        try:
            if nome_aba not in self.wb.sheetnames:
                print(f"❌ Aba '{nome_aba}' não encontrada!")
                return None
                
            ws = self.wb[nome_aba]
            
            if isinstance(data_alvo, str):
                data_obj = datetime.strptime(data_alvo, "%Y-%m-%d")
            else:
                data_obj = data_alvo
                
            meses_map = {
                "Jan": 1, "Fev": 2, "Mar": 3, "Abr": 4, 
                "Mai": 5, "Jun": 6, "Jul": 7, "Ago": 8, 
                "Set": 9, "Out": 10, "Nov": 11, "Dez": 12
            }
            
            # Define o intervalo baseado na aba
            if "1º" in nome_aba:
                dados = self.DATAS_POR_TRIMESTRE.get(1, {'inicio': 3, 'fim': 74})
            elif "2º" in nome_aba:
                dados = self.DATAS_POR_TRIMESTRE.get(2, {'inicio': 5, 'fim': 75})
            elif "3º" in nome_aba:
                dados = self.DATAS_POR_TRIMESTRE.get(3, {'inicio': 7, 'fim': 77})
            else:
                dados = {'inicio': 3, 'fim': 74}
            
            inicio, fim = dados['inicio'], dados['fim']
            print(f"   Buscando {nome_aba}: colunas {inicio} a {fim}")
            
            for col in range(inicio, fim + 1):
                celula = ws.cell(row=7, column=col)
                if isinstance(celula, MergedCell) or not celula.value:
                    continue
                    
                # O formato é VERTICAL com quebras de linha
                linhas = [l.strip() for l in str(celula.value).split('\n') if l.strip()]
                
                if len(linhas) >= 3:
                    try:
                        dezena = int(linhas[-1]) if linhas[-1].isdigit() else 0
                        unidade = int(linhas[-2]) if linhas[-2].isdigit() else 0
                        dia_planilha = (dezena * 10) + unidade
                        
                        letras_mes = "".join(reversed(linhas[:-2]))
                        mes_planilha = meses_map.get(letras_mes, 0)
                        
                        if dia_planilha == data_obj.day and mes_planilha == data_obj.month:
                            print(f"   ✅ Data encontrada na coluna {col} ({self._get_column_letter(col)})")
                            return col
                    except (ValueError, IndexError):
                        continue
                        
            print(f"   ❌ Data {data_alvo} não encontrada")
            return None
        except Exception as e:
            print(f"❌ Erro ao buscar coluna: {e}")
            return None

    def verificar_planilha_fechada(self):
        """
        Verifica se a planilha NÃO está aberta por nenhum processo.
        """
        if psutil is None:
            try:
                os.rename(self.caminho, self.caminho)
                return True
            except OSError:
                print("⚠️ Planilha bloqueada pelo sistema (provavelmente aberta).")
                return False
        
        try:
            caminho_abs = os.path.abspath(self.caminho)
            
            for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                try:
                    if proc.info['name'] and 'excel' in proc.info['name'].lower():
                        arquivos = proc.open_files()
                        if arquivos:
                            for arquivo in arquivos:
                                if caminho_abs.lower() in arquivo.path.lower():
                                    print(f"⚠️ Planilha aberta no processo: {proc.info['name']} (PID: {proc.info['pid']})")
                                    return False
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    continue
            
            print("✅ Planilha está fechada - pode salvar")
            return True
            
        except Exception as e:
            print(f"⚠️ Não foi possível verificar processos: {e}")
            try:
                os.rename(self.caminho, self.caminho)
                return True
            except OSError:
                print("⚠️ Planilha bloqueada pelo sistema (provavelmente aberta).")
                return False

    def carimbar_data_na_coluna(self, num_trimestre, coluna_alvo, data_obj):
        """
        Escreve a data na LINHA 7 da coluna especificada (formato VERTICAL)
        Esta função é usada apenas para a linha de frequência.
        
        Args:
            num_trimestre (int): Número do trimestre (1, 2, 3)
            coluna_alvo (int): Número da coluna onde escrever
            data_obj (datetime): Objeto de data
        """
        try:
            ws = self._obter_aba(num_trimestre)
            if not ws: 
                return False
            
            celula = ws.cell(row=7, column=coluna_alvo)
            if self._pode_escrever_na_celula(celula, coluna_alvo, 7):
                celula.value = self._formatar_data_linha7(data_obj)
                celula.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                celula.font = Font(size=7)
                return True
            return False
        except Exception as e:
            print(f"Erro ao carimbar data: {e}")
            return False

    def sincronizar_planejamento(self, num_trimestre, lista_planejamentos):
        """
        Sincroniza o planejamento de aulas na planilha.
        
        FAZ TRÊS COISAS:
        1. Escreve as DATAS na linha 7 (formato VERTICAL) - para frequência
        2. Escreve as DATAS nas colunas de planejamento (95/100/103) - formato HORIZONTAL (dd/mm/aaaa)
        3. Escreve os TEMAS nas colunas de planejamento (96/101/104)
        
        Args:
            num_trimestre (int): Número do trimestre (1, 2, 3)
            lista_planejamentos (list): Lista de tuplas (data, tema) ou dicionários
        """
        try:
            tri_str = str(num_trimestre)
            ws = self._obter_aba(tri_str)
            
            if not ws:
                print(f"❌ Aba {tri_str}º TRIMESTRE não encontrada!")
                return False

            if not lista_planejamentos:
                print("⚠️ Lista de planejamentos vazia")
                return False

            # Converte para lista padronizada de (data_obj, tema)
            planejamentos_ordenados = []
            for item in lista_planejamentos:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    data_obj = item[0]
                    tema = item[1]
                elif isinstance(item, dict):
                    data_obj = item.get('data')
                    tema = item.get('tema', '')
                else:
                    continue
                
                if isinstance(data_obj, str):
                    try:
                        data_obj = datetime.strptime(data_obj, "%Y-%m-%d")
                    except ValueError:
                        continue
                
                if isinstance(data_obj, datetime):
                    planejamentos_ordenados.append((data_obj, str(tema).upper()))
            
            if not planejamentos_ordenados:
                print("⚠️ Nenhuma data válida encontrada na lista")
                return False
                
            planejamentos_ordenados.sort(key=lambda x: x[0])

            # Obtém as colunas corretas para este trimestre
            cols_planejamento = self.PLANEJAMENTO_COLUNAS.get(num_trimestre)
            if not cols_planejamento:
                print(f"❌ Colunas de planejamento não definidas para o {num_trimestre}º trimestre")
                return False
            
            col_data_planej = cols_planejamento['data']  # 95, 100, ou 103
            col_tema_planej = cols_planejamento['tema']  # 96, 101, ou 104
            
            print(f"\n📅 Sincronizando {num_trimestre}º trimestre")
            print(f"   Coluna DATA planejamento: {col_data_planej}")
            print(f"   Coluna TEMA planejamento: {col_tema_planej}")

            # ========== 1. ESCREVE DATAS NA LINHA 7 (FREQUÊNCIA) ==========
            dados = self.DATAS_POR_TRIMESTRE.get(num_trimestre, {'inicio': 3, 'fim': 73})
            coluna_atual = dados['inicio']
            fim = dados['fim']
            
            print(f"\n📅 Escrevendo datas na linha 7 (formato vertical) - colunas {coluna_atual} a {fim}")
            
            # Limpa as datas da linha 7
            for col in range(coluna_atual, fim + 1):
                celula = ws.cell(row=7, column=col)
                if self._pode_escrever_na_celula(celula, col, 7):
                    celula.value = None

            # Escreve as novas datas na linha 7
            col_atual = coluna_atual
            for data_obj, tema in planejamentos_ordenados:
                if col_atual > fim:
                    print(f"⚠️ Aviso: {len(planejamentos_ordenados)} aulas excedem o espaço disponível")
                    break
                celula = ws.cell(row=7, column=col_atual)
                if self._pode_escrever_na_celula(celula, col_atual, 7):
                    celula.value = self._formatar_data_linha7(data_obj)
                    celula.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    celula.font = Font(size=7)
                    print(f"  📍 Data {data_obj.strftime('%d/%m/%Y')} -> linha 7, coluna {col_atual}")
                col_atual += 1

            # ========== 2. LIMPA E ESCREVE DATAS NAS COLUNAS DE PLANEJAMENTO ==========
            print(f"\n📅 Escrevendo datas na coluna {col_data_planej} (formato horizontal)...")
            
            # LIMPA TODA A COLUNA DE DATA DO PLANEJAMENTO (linhas 9 a 100)
            for linha in range(9, 101):
                celula = ws.cell(row=linha, column=col_data_planej)
                if self._pode_escrever_na_celula(celula, col_data_planej, linha):
                    if celula.value is not None:
                        print(f"  🗑️ Limpando célula DATA: col{col_data_planej}, linha{linha} (era: {celula.value})")
                    celula.value = None
            
            # LIMPA TODA A COLUNA DE TEMA DO PLANEJAMENTO (linhas 9 a 100)
            print(f"\n📝 Escrevendo temas na coluna {col_tema_planej}...")
            
            for linha in range(9, 101):
                celula = ws.cell(row=linha, column=col_tema_planej)
                if self._pode_escrever_na_celula(celula, col_tema_planej, linha):
                    if celula.value is not None:
                        print(f"  🗑️ Limpando célula TEMA: col{col_tema_planej}, linha{linha} (era: {celula.value})")
                    celula.value = None
            
            # ========== 3. ESCREVE AS NOVAS DATAS E TEMAS ==========
            print(f"\n✍️ Escrevendo novos dados...")
            
            linha_atual = 9
            for data_obj, tema in planejamentos_ordenados:
                if linha_atual > 100:
                    break
                
                # Escreve DATA
                celula_data = ws.cell(row=linha_atual, column=col_data_planej)
                if self._pode_escrever_na_celula(celula_data, col_data_planej, linha_atual):
                    data_horizontal = self._formatar_data_horizontal(data_obj, incluir_ano=True)
                    celula_data.value = data_horizontal
                    celula_data.alignment = Alignment(horizontal='center')
                    print(f"  📅 Data: {data_horizontal} (col{col_data_planej}, linha{linha_atual})")
                
                # Escreve TEMA
                celula_tema = ws.cell(row=linha_atual, column=col_tema_planej)
                if self._pode_escrever_na_celula(celula_tema, col_tema_planej, linha_atual):
                    celula_tema.value = tema
                    celula_tema.alignment = Alignment(horizontal='left')
                    print(f"  📖 Tema: {tema[:50]}... (col{col_tema_planej}, linha{linha_atual})")
                
                linha_atual += 1

            # ========== 4. LIMPA AS LINHAS RESTANTES (se houver) ==========
            # Isso garante que não fiquem dados órfãos de execuções anteriores
            while linha_atual <= 100:
                celula_data = ws.cell(row=linha_atual, column=col_data_planej)
                if self._pode_escrever_na_celula(celula_data, col_data_planej, linha_atual):
                    if celula_data.value is not None:
                        print(f"  🗑️ Limpando célula DATA órfã: col{col_data_planej}, linha{linha_atual}")
                    celula_data.value = None
                
                celula_tema = ws.cell(row=linha_atual, column=col_tema_planej)
                if self._pode_escrever_na_celula(celula_tema, col_tema_planej, linha_atual):
                    if celula_tema.value is not None:
                        print(f"  🗑️ Limpando célula TEMA órfã: col{col_tema_planej}, linha{linha_atual}")
                    celula_tema.value = None
                linha_atual += 1

            print(f"\n✅ Planejamento do {num_trimestre}º trimestre concluído: {len(planejamentos_ordenados)} aulas")
            return True
            
        except Exception as e:
            print(f"❌ Erro no sincronizar_planejamento: {e}")
            import traceback
            traceback.print_exc()
            return False


    def marcar_faltas_por_data(self, num_trimestre, data_aula, lista_faltosos):
        """
        Marca faltas (F) na coluna correspondente à data especificada.
        
        Args:
            num_trimestre (int): Número do trimestre (1, 2, 3)
            data_aula (str ou datetime): Data da aula
            lista_faltosos (list): Lista de nomes dos alunos faltosos
        """
        try:
            nome_aba = f"{num_trimestre}º TRIMESTRE"
            coluna_data = self.buscar_coluna_por_data(nome_aba, data_aula)
            
            if coluna_data is None:
                print(f"❌ Data {data_aula} não encontrada!")
                return False
            
            ws = self._obter_aba(num_trimestre)
            if not ws:
                return False
            
            lancamentos = 0
            for nome_faltoso in lista_faltosos:
                nome_upper = nome_faltoso.strip().upper()
                if nome_upper in self.cache_alunos:
                    linha = self.cache_alunos[nome_upper]
                    celula_falta = ws.cell(row=linha, column=coluna_data)
                    if self._pode_escrever_na_celula(celula_falta, coluna_data, linha):
                        celula_falta.value = "F"
                        celula_falta.alignment = Alignment(horizontal='center')
                        lancamentos += 1
                        print(f"  ✅ Falta marcada: {nome_faltoso}")
                else:
                    print(f"  ⚠️ Aluno não encontrado no cache: {nome_faltoso}")
            
            if lancamentos == 0:
                print("ℹ️ Nenhuma falta registrada (todos presentes)")
            else:
                print(f"✅ {lancamentos} faltas marcadas na data {data_aula}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao marcar faltas: {e}")
            return False

    def lancar_notas_com_validacao(self, num_trimestre, coluna_alvo, lista_notas):
        """
        Lança notas na coluna especificada, validando se a coluna não é travada.
        
        Args:
            num_trimestre (int): Número do trimestre (1, 2, 3)
            coluna_alvo (int): Número da coluna onde escrever as notas
            lista_notas (list): Lista de tuplas (nome_aluno, nota)
        """
        try:
            tri_str = str(num_trimestre)
            
            # Verifica coluna travada
            if coluna_alvo in self.COLUNAS_TRAVADAS.get(num_trimestre, []):
                print(f"❌ Coluna {self._get_column_letter(coluna_alvo)} é travada (fórmula)!")
                return False
            
            ws = self._obter_aba(tri_str)
            if not ws:
                print(f"❌ Aba {tri_str}º TRIMESTRE não encontrada!")
                return False
            
            if not lista_notas:
                print(f"⚠️ Lista de notas vazia")
                return False
            
            print(f"📊 Processando {len(lista_notas)} notas para coluna {coluna_alvo}")
            
            # Limpa valores anteriores (apenas células editáveis)
            for linha in self.cache_alunos.values():
                celula = ws.cell(row=linha, column=coluna_alvo)
                if self._pode_escrever_na_celula(celula, coluna_alvo, linha):
                    celula.value = None
            
            sucessos = 0
            for nome_aluno, nota_valor in lista_notas:
                nome_upper = nome_aluno.strip().upper()
                if nome_upper in self.cache_alunos:
                    linha = self.cache_alunos[nome_upper]
                    celula = ws.cell(row=linha, column=coluna_alvo)
                    if self._pode_escrever_na_celula(celula, coluna_alvo, linha):
                        try:
                            celula.value = float(nota_valor)
                            celula.alignment = Alignment(horizontal='center')
                            celula.font = Font(bold=True)
                            sucessos += 1
                            print(f"  ✅ Nota {nota_valor} para {nome_aluno}")
                        except (ValueError, TypeError):
                            celula.value = 0
                            sucessos += 1
                    else:
                        print(f"  ⚠️ {nome_aluno}: célula com fórmula ou travada!")
                else:
                    print(f"  ❌ {nome_aluno}: não encontrado no cache!")
            
            print(f"✅ Sincronização concluída: {sucessos} notas gravadas")
            return sucessos > 0
            
        except Exception as e:
            print(f"❌ Erro ao lançar notas: {e}")
            import traceback
            traceback.print_exc()
            return False

    def salvar(self):
        """Salva as alterações no arquivo Excel"""
        try:
            self.wb.save(self.caminho)
            print("✅ Planilha salva com sucesso!")
            return True
        except PermissionError:
            print("❌ Permissão negada! Planilha pode estar aberta.")
            return False
        except Exception as e:
            print(f"❌ Erro ao salvar: {e}")
            return False
            
    def fechar(self):
        """Fecha a planilha (libera recursos)"""
        try:
            self.wb.close()
        except:
            pass